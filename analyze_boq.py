import openpyxl
from openpyxl.utils import get_column_letter
import json

# Load the workbook
wb = openpyxl.load_workbook('Template BOQ.xlsx')
ws = wb.active

print("=" * 80)
print(f"SHEET NAME: {ws.title}")
print(f"MAX ROW: {ws.max_row}")
print(f"MAX COLUMN: {ws.max_column}")
print("=" * 80)
print()

# Extract all content
data = []
for row_idx in range(1, min(ws.max_row + 1, 50)):  # First 50 rows
    row_data = []
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        
        # Get cell info
        value = cell.value if cell.value is not None else ""
        
        # Get merge info
        is_merged = False
        merge_info = ""
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                is_merged = True
                merge_info = f" [MERGED: {merged_range}]"
                break
        
        # Get fill color
        fill_color = ""
        if cell.fill and cell.fill.start_color:
            if cell.fill.start_color.rgb and cell.fill.start_color.rgb != '00000000':
                fill_color = f" [FILL: {cell.fill.start_color.rgb}]"
        
        # Get border
        has_border = False
        if cell.border:
            if any([cell.border.left.style, cell.border.right.style, 
                   cell.border.top.style, cell.border.bottom.style]):
                has_border = True
        
        border_info = " [BORDER]" if has_border else ""
        
        row_data.append(f"{value}{merge_info}{fill_color}{border_info}")
    
    data.append(row_data)
    print(f"ROW {row_idx:2d}: {row_data}")

print()
print("=" * 80)
print("MERGED CELLS:")
for merged_range in ws.merged_cells.ranges:
    print(f"  {merged_range}")

print()
print("=" * 80)
print("ANALYSIS COMPLETE")
